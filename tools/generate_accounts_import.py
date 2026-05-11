"""
Excel → Supabase SQL import generator
মাতবাখ-মাদরাসা-৪৭-৪৮.xlsx  →  mdr_account_* tables
"""
import openpyxl
import sys, io, re, random, string, unicodedata
from datetime import date, datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

EXCEL_PATH = r'C:\Users\sanim\Downloads\মাতবাখ-মাদরাসা-৪৭-৪৮.xlsx'
SOURCE_FILE = 'মাতবাখ-মাদরাসা-৪৭-৪৮.xlsx'
ACADEMIC_YEAR_LABEL = '১৪৪৭/৪৮'

# ── Month normalization ─────────────────────────────────────────────────────
MONTH_ALIAS = {
    'রামাযান':'রমজান','রমাযান':'রমজান','রমযান':'রমজান',
    'রামজান':'রমজান',
    'যিলকদ':'জিলকদ','যিলক্বদ':'জিলকদ','যুলকদ':'জিলকদ',
    'জিলকাদ':'জিলকদ','জিলক্বদ':'জিলকদ','জুলকদ':'জিলকদ','জিলকাদ':'জিলকদ',
    'যিলহাজ':'জিলহজ','যিলহজ':'জিলহজ','যুলহজ':'জিলহজ',
    'জিলহাজ':'জিলহজ','জুলহজ':'জিলহজ','জিলহাজ্ব':'জিলহজ',
    'মুহার্‌রম':'মুহাররম','মুহার্রম':'মুহাররম','মুহর্‌রম':'মুহাররম',
    'রবিউল আঃ':'রবিউল আউয়াল','রবিউল আওয়াল':'রবিউল আউয়াল',
    'রবিউল আঃ':'রবিউল আউয়াল',
    'রবিউস সানিঃ':'রবিউস সানি','রবিউস সানী':'রবিউস সানি',
    'জুমাদাল আখেরা':'জুমাদাল উখরা','জুমাদাল উখরা':'জুমাদাল উখরা',
    'জুমা:আখেরা':'জুমাদাল উখরা','জুমাঃআঃ':'জুমাদাল উখরা',
    'জমাদাল উলা':'জুমাদাল উলা','জমাদাল উখরা':'জুমাদাল উখরা',
    'জমাদিউল আউয়াল':'জুমাদাল উলা','জমাদিউস সানি':'জুমাদাল উখরা',
    'জুমাঃউঃ':'জুমাদাল উলা','জুমা:উলা':'জুমাদাল উলা',
    'জুমাদাল উলা':'জুমাদাল উলা',
    'রজব':'রজব','শাবান':'শাবান','শাওয়াল':'শাওয়াল',
    'সফর':'সফর','মুহাররম':'মুহাররম',
    'রবিউল আউয়াল':'রবিউল আউয়াল','রবিউস সানি':'রবিউস সানি',
    'জুমাদাল উলা':'জুমাদাল উলা',
    'রমজান':'রমজান','জিলকদ':'জিলকদ','জিলহজ':'জিলহজ',
    'জিলকদ':'জিলকদ',
}

MONTH_NUM = {
    'মুহাররম':1,'সফর':2,'রবিউল আউয়াল':3,'রবিউস সানি':4,
    'জুমাদাল উলা':5,'জুমাদাল উখরা':6,'রজব':7,'শাবান':8,
    'রমজান':9,'শাওয়াল':10,'জিলকদ':11,'জিলহজ':12,
}

ACCOUNT_MAP = {
    'মাতবাখ':'matbakh','matbakh':'matbakh',
    'মাদরাসা':'madrasa','madrasa':'madrasa',
    'তামিরাত':'tamirat','তামীরাত':'tamirat','tamirat':'tamirat',
    'সাধারণ':'general','general':'general',
    'বাসার হিসাব':'general',
}

# Approximate Gregorian→Hijri ranges for 1447-48
# Format: (start_date, end_date_inclusive, hijri_month, hijri_year)
GREG_HIJRI_RANGES = [
    (date(2025, 7, 7),  date(2025, 8, 5),  'মুহাররম', 1447),
    (date(2025, 8, 6),  date(2025, 9, 3),  'সফর', 1447),
    (date(2025, 9, 4),  date(2025, 10, 3), 'রবিউল আউয়াল', 1447),
    (date(2025, 10, 4), date(2025, 11, 1), 'রবিউস সানি', 1447),
    (date(2025, 11, 2), date(2025, 11, 30),'জুমাদাল উলা', 1447),
    (date(2025, 12, 1), date(2025, 12, 29),'জুমাদাল উখরা', 1447),
    (date(2025, 12, 30),date(2026, 1, 27), 'রজব', 1447),
    (date(2026, 1, 28), date(2026, 2, 25), 'শাবান', 1447),
    (date(2026, 2, 26), date(2026, 3, 28), 'রমজান', 1447),
    (date(2026, 3, 29), date(2026, 4, 26), 'শাওয়াল', 1447),
    (date(2026, 4, 27), date(2026, 5, 25), 'জিলকদ', 1447),
    (date(2026, 5, 26), date(2026, 6, 24), 'জিলহজ', 1447),
    (date(2026, 6, 25), date(2026, 7, 23), 'মুহাররম', 1448),
    (date(2026, 7, 24), date(2026, 8, 21), 'সফর', 1448),
]

def greg_to_hijri_full(d):
    """Gregorian date → (hijri_month, hijri_year, hijri_day ১–৩০)। রেঞ্জের প্রথম দিন ≈ হিজরী ১।"""
    if d is None:
        return None, None, None
    if isinstance(d, datetime):
        d = d.date()
    if not isinstance(d, date):
        return None, None, None
    for start, end, month, year in GREG_HIJRI_RANGES:
        if start <= d <= end:
            day_h = min(30, max(1, (d - start).days + 1))
            return month, year, day_h
    return None, None, None

def greg_to_hijri(d):
    """Gregorian date → (hijri_month, hijri_year) — পুরনো কলের জন্য"""
    m, y, _ = greg_to_hijri_full(d)
    return m, y

def nfc(s):
    """NFC normalize + strip invisible chars — handles Excel's NFD Bengali chars"""
    return unicodedata.normalize('NFC', str(s)).replace('\u200c','').replace('\u200d','').strip()

def to_en_digits(s):
    """বাংলা/আরবি/ফার্সি অঙ্ক → ASCII অঙ্ক শুধু (সংখ্যা নয় এমন অক্ষর বাদ)"""
    if s is None:
        return ''
    out = []
    for ch in str(s).strip():
        if '\u09e6' <= ch <= '\u09ef':
            out.append(str(ord(ch) - ord('\u09e6')))
        elif '\u0660' <= ch <= '\u0669':
            out.append(str(ord(ch) - ord('\u0660')))
        elif '\u06f0' <= ch <= '\u06f9':
            out.append(str(ord(ch) - ord('\u06f0')))
        elif ch.isdigit():
            out.append(ch)
    return ''.join(out)

def parse_hijri_day_cell(v):
    """এক্সেল দিন সেল: ইংরেজি/বাংলা সংখ্যা, অথবা ৫/১১ টাইপ — দিন ১–৩০"""
    if v is None or v == '':
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            i = int(v)
            return i if 1 <= i <= 30 else None
        except (ValueError, TypeError, OverflowError):
            return None
    raw = nfc(str(v).strip())
    if not raw:
        return None
    for sep in ['/', '-', '।']:
        if sep in raw:
            bits = [b.strip() for b in raw.replace('।', '.').split(sep) if str(b).strip()]
            for b in reversed(bits):
                s2 = to_en_digits(b)
                if not s2 or len(s2) > 2:
                    continue
                try:
                    i = int(s2)
                    if 1 <= i <= 30:
                        return i
                except ValueError:
                    pass
    s = to_en_digits(raw)
    if not s:
        return None
    try:
        i = int(s)
        if 1 <= i <= 30:
            return i
    except ValueError:
        pass
    return None

# Build NFC-normalized alias dict so lookup works regardless of source encoding
_MONTH_ALIAS_NFC = {nfc(k): v for k, v in MONTH_ALIAS.items()}
_MONTH_NUM_NFC   = {nfc(k): v for k, v in MONTH_NUM.items()}

def norm_month(m):
    if not m or str(m).strip() == '':
        return None
    s = nfc(str(m).strip())
    s = s.rstrip(':;।')
    # Try alias lookup first (handles variants like জিলকাদ → জিলকদ)
    if s in _MONTH_ALIAS_NFC:
        return _MONTH_ALIAS_NFC[s]
    # Already a canonical month name?
    if s in _MONTH_NUM_NFC:
        return s
    return None

def hijri_year_from_month(month_name):
    """Academic year 1447/48: months 9-12 → 1447, months 1-8 → 1448"""
    n = MONTH_NUM.get(month_name, 0)
    if n == 0:
        return '1447'  # default
    return '1447' if n >= 9 else '1448'

def num_val(v):
    if v is None:
        return None
    try:
        f = float(str(v).replace(',',''))
        return f if f != 0 else None
    except:
        return None

def str_val(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def sql_str(v):
    if v is None:
        return 'NULL'
    s = str(v).replace("'", "''")
    return f"'{s}'"

def sql_num(v):
    if v is None:
        return 'NULL'
    try:
        f = float(v)
        return str(round(f, 4))
    except:
        return 'NULL'

def sql_int(v):
    if v is None:
        return 'NULL'
    try:
        return str(int(float(str(v))))
    except:
        return 'NULL'

def sql_date(v):
    if v is None:
        return 'NULL'
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d')}'"
    if isinstance(v, date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    return 'NULL'

_id_counter = [0]
def gen_id(prefix):
    _id_counter[0] += 1
    return f"{prefix}{_id_counter[0]:05d}"

# ── Load workbook ──────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

rows_income = []
rows_expense = []
rows_due = []
rows_due_payment = []

# ═══════════════════════════════════════════════════════════════════════════
# 1. অর্থ গ্রহণ → mdr_account_incomes
# Col: A=মাস, B=day, C=পরিমাণ, D=মন্তব্য, E=খাত
# ═══════════════════════════════════════════════════════════════════════════
ws = wb['অর্থ গ্রহণ']
for rno, row in enumerate(ws.iter_rows(min_row=11, values_only=True), start=11):
    month_raw = str_val(row[0])
    day_val   = str_val(row[1])
    amount    = num_val(row[2])
    note      = str_val(row[3])
    acct_raw  = str_val(row[4])
    if not month_raw and not amount:
        continue
    month = norm_month(month_raw)
    if not month or not amount:
        continue
    acct = ACCOUNT_MAP.get(str(acct_raw or '').strip(), 'general')
    hijri_year = hijri_year_from_month(month)
    day_int = parse_hijri_day_cell(day_val)
    rows_income.append({
        'id': gen_id('inc-47-'),
        'account': acct,
        'hijri_year': hijri_year,
        'month': month,
        'day': day_int,
        'gregorian_date': None,
        'amount': amount,
        'note': note or '',
        'source_sheet': 'অর্থ গ্রহণ',
        'source_row': rno,
    })

# ═══════════════════════════════════════════════════════════════════════════
# 2. মাতবাখ → mdr_account_expenses (account='matbakh')
# Col: A=মাস, B=day, C=শাখা খাত, D=বিবরণ, E=পরিমাণ, F=মূল্য, G=মোট,
#      H=সরবরাহকারী, I=রশিদ
# ═══════════════════════════════════════════════════════════════════════════
ws = wb['মাতবাখ']
for rno, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
    month_raw = str_val(row[0])
    day_raw   = row[1]
    category  = str_val(row[2])
    desc      = str_val(row[3])
    qty       = num_val(row[4])
    unit_price= num_val(row[5])
    amount    = num_val(row[6])
    supplier  = str_val(row[7])
    receipt   = str_val(row[8])
    # Skip empty rows
    if not month_raw and not amount and not desc:
        continue
    month = norm_month(month_raw)
    if not month or not amount:
        continue
    day_int = parse_hijri_day_cell(day_raw)
    hijri_year = hijri_year_from_month(month)
    rows_expense.append({
        'id': gen_id('exp-mb-'),
        'account': 'matbakh',
        'project': None,
        'hijri_year': hijri_year,
        'month': month,
        'day': day_int,
        'gregorian_date': None,
        'category': category,
        'description': desc or '',
        'quantity': qty,
        'unit': None,
        'unit_price': unit_price,
        'amount': amount,
        'supplier': supplier,
        'receipt_no': receipt,
        'payment_method': None,
        'source_sheet': 'মাতবাখ',
        'source_row': rno,
    })

# ═══════════════════════════════════════════════════════════════════════════
# 3. মাদরাসা → mdr_account_expenses (account='madrasa')
# Col: A=serial, B=মাস, C=day, D=শাখা খাত, E=বিবরণ, F=পরিমাণ,
#      G=মূল্য, H=মোট, I=সরবরাহকারী, J=রশীদ
# ═══════════════════════════════════════════════════════════════════════════
ws = wb['মাদরাসা']
for rno, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    month_raw = str_val(row[1])
    day_raw   = row[2]
    category  = str_val(row[3])
    desc      = str_val(row[4])
    qty       = num_val(row[5])
    unit_price= num_val(row[6])
    amount    = num_val(row[7])
    supplier  = str_val(row[8])
    receipt   = str_val(row[9])
    if not month_raw and not amount and not desc:
        continue
    month = norm_month(month_raw)
    if not month or not amount:
        continue
    day_int = parse_hijri_day_cell(day_raw)
    hijri_year = hijri_year_from_month(month)
    rows_expense.append({
        'id': gen_id('exp-md-'),
        'account': 'madrasa',
        'project': None,
        'hijri_year': hijri_year,
        'month': month,
        'day': day_int,
        'gregorian_date': None,
        'category': category,
        'description': desc or '',
        'quantity': qty,
        'unit': None,
        'unit_price': unit_price,
        'amount': amount,
        'supplier': supplier,
        'receipt_no': receipt,
        'payment_method': None,
        'source_sheet': 'মাদরাসা',
        'source_row': rno,
    })

# ═══════════════════════════════════════════════════════════════════════════
# 4. তামীরাত-খরচ → mdr_account_expenses (account='tamirat')
# Col: A=মাস (Hijri), B=দিন, C=খাত/project, D=বিবরণ, E=পরিমাণ,
#      F=মাপ/unit, G=মূল্য, H=খরচ/amount, I=বাকি, J=সরবরাহকারী
# (format updated — now same as matbakh/madrasa sheets)
# ═══════════════════════════════════════════════════════════════════════════
ws = wb['তামীরাত-খরচ']
for rno, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
    month_raw = str_val(row[0])
    day_raw   = row[1]
    project   = str_val(row[2])
    desc      = str_val(row[3])
    qty       = num_val(row[4])
    unit      = str_val(row[5])
    unit_price= num_val(row[6])
    amount    = num_val(row[7])
    supplier  = str_val(row[9])
    if not amount:
        continue
    month = norm_month(month_raw)
    if not month:
        continue
    day = parse_hijri_day_cell(day_raw)
    hyear = hijri_year_from_month(month)
    rows_expense.append({
        'id': gen_id('exp-tm-'),
        'account': 'tamirat',
        'project': project,
        'hijri_year': str(hyear),
        'month': month,
        'day': day,
        'gregorian_date': None,
        'category': project,
        'description': desc or '',
        'quantity': qty,
        'unit': unit,
        'unit_price': unit_price,
        'amount': amount,
        'supplier': supplier,
        'receipt_no': None,
        'payment_method': None,
        'source_sheet': 'তামীরাত-খরচ',
        'source_row': rno,
    })

# ═══════════════════════════════════════════════════════════════════════════
# 5. বাসার হিসাব → mdr_account_expenses (account='general')
# Col: A=মাস, B=day, C=শাখা খাত, D=বিবরণ, E=পরিমাণ, F=মূল্য, G=মোট,
#      H=সরবরাহকারী, I=রশীদ
# ═══════════════════════════════════════════════════════════════════════════
ws = wb['বাসার হিসাব']
for rno, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
    month_raw = str_val(row[0])
    day_raw   = row[1]
    category  = str_val(row[2])
    desc      = str_val(row[3])
    qty       = num_val(row[4])
    unit_price= num_val(row[5])
    amount    = num_val(row[6])
    supplier  = str_val(row[7])
    receipt   = str_val(row[8])
    if not month_raw and not amount and not desc:
        continue
    month = norm_month(month_raw)
    if not month or not amount:
        continue
    day_int = parse_hijri_day_cell(day_raw)
    hijri_year = hijri_year_from_month(month)
    rows_expense.append({
        'id': gen_id('exp-bs-'),
        'account': 'general',
        'project': None,
        'hijri_year': hijri_year,
        'month': month,
        'day': day_int,
        'gregorian_date': None,
        'category': category,
        'description': desc or '',
        'quantity': qty,
        'unit': None,
        'unit_price': unit_price,
        'amount': amount,
        'supplier': supplier,
        'receipt_no': receipt,
        'payment_method': None,
        'source_sheet': 'বাসার হিসাব',
        'source_row': rno,
    })

# ═══════════════════════════════════════════════════════════════════════════
# 6. বিভিন্ন স্থানে অর্থ → mdr_account_expenses (account='general')
# Col: A=তারিখ (Gregorian), B=নাম/desc, C=পরিমাণ, D=মাধ্যম, E=খাত
# ═══════════════════════════════════════════════════════════════════════════
ws = wb['বিভিন্ন স্থানে অর্থ ']
for rno, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
    greg_date = row[0]
    desc      = str_val(row[1])
    amount    = num_val(row[2])
    pay_meth  = str_val(row[3])
    if not amount and not desc:
        continue
    if not amount:
        continue
    greg = None
    if isinstance(greg_date, (date, datetime)):
        greg = greg_date if isinstance(greg_date, date) else greg_date.date()
    month, hyear, hday = greg_to_hijri_full(greg)
    if not month:
        month = 'জিলকদ'
        hyear = 1447
        hday = hday or 1
    rows_expense.append({
        'id': gen_id('exp-bf-'),
        'account': 'general',
        'project': 'বিভিন্ন স্থানে অর্থ',
        'hijri_year': str(hyear),
        'month': month,
        'day': hday,
        'gregorian_date': greg,
        'category': 'বিভিন্ন স্থানে অর্থ',
        'description': desc or '',
        'quantity': None,
        'unit': None,
        'unit_price': None,
        'amount': amount,
        'supplier': None,
        'receipt_no': None,
        'payment_method': pay_meth,
        'source_sheet': 'বিভিন্ন স্থানে অর্থ ',
        'source_row': rno,
    })

# ═══════════════════════════════════════════════════════════════════════════
# 7. বাকির হিসাব → dues + due_payments
# Top section (rows 2-9): supplier summary → mdr_account_dues
# Bottom section (rows 12+): payment entries → mdr_account_due_payments
# ═══════════════════════════════════════════════════════════════════════════
ws = wb['বাকির হিসাব']
# Top left: বিক্রেতা ১-৭ (matbakh/madrasa suppliers)
SUPPLIER_ACCOUNT = {
    'সাগর': 'matbakh', 'আল মদিনা': 'matbakh', 'গ্যাস': 'matbakh',
    'ডিম': 'matbakh', 'বেলাল': 'matbakh', 'লাকড়ি': 'matbakh',
    'জমিনের সবজী': 'matbakh',
    'আনোয়ার': 'tamirat', 'বিশাল থাই': 'tamirat', 'কামাল সিমেন্ট': 'tamirat',
    'বন্ধু ট্রেডার্স': 'tamirat', 'ঊষা স্যানেটারি': 'tamirat',
    'বেলাল আলীপুর': 'tamirat', 'মা হার্ডওয়ার': 'tamirat',
    'কাঠ মিস্ত্রি': 'tamirat',
}
# rows 2-9: col B=vendor_no, C=supplier, D=total, E=paid, F=due
due_rows_raw = list(ws.iter_rows(min_row=2, max_row=9, values_only=True))
# rows 2-9: col I=vendor_no, J=supplier, K=total, L=paid (not present), M=?
# Actually from the analysis: col H=vendor8, I=supplier8, J=total, K=paid, L=due
all_due_rows = []
for row in due_rows_raw:
    # Left side (B=no, C=name, D=total, E=paid, F=due)
    sup_left  = str_val(row[2])
    tot_left  = num_val(row[3])
    paid_left = num_val(row[4])
    due_left  = num_val(row[5])
    if sup_left and tot_left is not None:
        all_due_rows.append((sup_left, tot_left, paid_left, due_left))
    # Right side: col H=vendor_no(row[7]), I=supplier(row[8]), J=total(row[9]), K=paid(row[10]), L=due(row[11])
    if len(row) >= 10:
        sup_right  = str_val(row[8])
        tot_right  = num_val(row[9])
        paid_right = num_val(row[10])
        due_right  = num_val(row[11]) if len(row) > 11 else None
        if sup_right and tot_right is not None:
            all_due_rows.append((sup_right, tot_right, paid_right, due_right))

for sup, total, paid, due in all_due_rows:
    if not sup:
        continue
    acct = SUPPLIER_ACCOUNT.get(sup, 'matbakh')
    total_v = total or 0
    paid_v  = paid or 0
    due_v   = due if due is not None else (total_v - paid_v)
    rows_due.append({
        'id': gen_id('due-47-'),
        'account': acct,
        'supplier': sup,
        'total': total_v,
        'paid': paid_v,
        'due': due_v,
        'source_sheet': 'বাকির হিসাব',
    })

# Payment rows (from row 12 onwards): B=date, C=amount, D=supplier, E=receipt
for rno, row in enumerate(ws.iter_rows(min_row=12, values_only=True), start=12):
    greg_date = row[1]
    amount    = num_val(row[2])
    supplier  = str_val(row[3])
    receipt   = str_val(row[4])
    if not amount or not supplier:
        continue
    greg = None
    if isinstance(greg_date, (date, datetime)):
        greg = greg_date if isinstance(greg_date, date) else greg_date.date()
    month, hyear, hday = greg_to_hijri_full(greg)
    if not month:
        month = 'শাওয়াল'
        hyear = 1447
        hday = hday or 1
    acct = SUPPLIER_ACCOUNT.get(supplier, 'matbakh')
    # Find matching due_id
    due_id = None
    for d in rows_due:
        if d['supplier'] == supplier and d['account'] == acct:
            due_id = d['id']
            break
    rows_due_payment.append({
        'id': gen_id('dpay-47-'),
        'due_id': due_id,
        'account': acct,
        'supplier': supplier,
        'hijri_year': str(hyear),
        'month': month,
        'day': hday,
        'gregorian_date': greg,
        'amount': amount,
        'receipt_no': receipt,
        'source_sheet': 'বাকির হিসাব',
        'source_row': rno,
    })

# ═══════════════════════════════════════════════════════════════════════════
# Generate SQL
# ═══════════════════════════════════════════════════════════════════════════
out = []
out.append('-- ═══════════════════════════════════════════════════════════════')
out.append(f'-- মাতবাখ-মাদরাসা হিসাব import — {ACADEMIC_YEAR_LABEL} হিজরী')
out.append('-- Generated by tools/generate_accounts_import.py')
out.append('-- WARNING: এই SQL চালানোর আগে user-এর explicit approval নিতে হবে!')
out.append('-- ═══════════════════════════════════════════════════════════════')
out.append('')

# DELETE section
out.append('-- ── পুরনো ডেটা মুছে ফেলা ─────────────────────────────────────')
out.append('-- মোট বর্তমান রেকর্ড:')
out.append('--   incomes: 18, expenses: 715, dues: 16, due_payments: 14')
out.append('DELETE FROM public.mdr_account_due_payments;')
out.append('DELETE FROM public.mdr_account_dues;')
out.append('DELETE FROM public.mdr_account_expenses;')
out.append('DELETE FROM public.mdr_account_incomes;')
out.append('')

# Summary comment
out.append(f'-- ── নতুন ডেটা ──────────────────────────────────────────────────')
out.append(f'-- Income rows  : {len(rows_income)}')
out.append(f'-- Expense rows : {len(rows_expense)}')
out.append(f'--   matbakh    : {sum(1 for r in rows_expense if r["account"]=="matbakh")}')
out.append(f'--   madrasa    : {sum(1 for r in rows_expense if r["account"]=="madrasa")}')
out.append(f'--   tamirat    : {sum(1 for r in rows_expense if r["account"]=="tamirat")}')
out.append(f'--   general    : {sum(1 for r in rows_expense if r["account"]=="general")}')
out.append(f'-- Due rows     : {len(rows_due)}')
out.append(f'-- Due payment  : {len(rows_due_payment)}')
out.append('')

# Income inserts
if rows_income:
    out.append('-- ── mdr_account_incomes ────────────────────────────────────────')
    out.append('INSERT INTO public.mdr_account_incomes')
    out.append('  (id, account, hijri_year, month, day, gregorian_date, amount, note, source_file, source_sheet, source_row)')
    out.append('VALUES')
    vals = []
    for r in rows_income:
        vals.append(
            f"  ({sql_str(r['id'])}, {sql_str(r['account'])}, {sql_str(r['hijri_year'])}, "
            f"{sql_str(r['month'])}, {sql_int(r['day'])}, {sql_date(r['gregorian_date'])}, "
            f"{sql_num(r['amount'])}, {sql_str(r['note'])}, {sql_str(SOURCE_FILE)}, "
            f"{sql_str(r['source_sheet'])}, {sql_int(r['source_row'])})"
        )
    out.append(',\n'.join(vals) + ';')
    out.append('')

# Expense inserts — batch in chunks of 500 to avoid huge single statements
CHUNK = 500
for chunk_start in range(0, len(rows_expense), CHUNK):
    chunk = rows_expense[chunk_start:chunk_start + CHUNK]
    out.append(f'-- ── mdr_account_expenses (rows {chunk_start+1}–{chunk_start+len(chunk)}) ──')
    out.append('INSERT INTO public.mdr_account_expenses')
    out.append('  (id, account, project, hijri_year, month, day, gregorian_date, category, description,')
    out.append('   quantity, unit, unit_price, amount, supplier, receipt_no, payment_method,')
    out.append('   source_file, source_sheet, source_row)')
    out.append('VALUES')
    vals = []
    for r in chunk:
        vals.append(
            f"  ({sql_str(r['id'])}, {sql_str(r['account'])}, {sql_str(r['project'])}, "
            f"{sql_str(r['hijri_year'])}, {sql_str(r['month'])}, {sql_int(r['day'])}, "
            f"{sql_date(r['gregorian_date'])}, {sql_str(r['category'])}, {sql_str(r['description'])}, "
            f"{sql_num(r['quantity'])}, {sql_str(r['unit'])}, {sql_num(r['unit_price'])}, "
            f"{sql_num(r['amount'])}, {sql_str(r['supplier'])}, {sql_str(r['receipt_no'])}, "
            f"{sql_str(r['payment_method'])}, {sql_str(SOURCE_FILE)}, "
            f"{sql_str(r['source_sheet'])}, {sql_int(r['source_row'])})"
        )
    out.append(',\n'.join(vals) + ';')
    out.append('')

# Dues inserts
if rows_due:
    out.append('-- ── mdr_account_dues ───────────────────────────────────────────')
    out.append('INSERT INTO public.mdr_account_dues')
    out.append('  (id, account, supplier, total, paid, due, source_file, source_sheet)')
    out.append('VALUES')
    vals = []
    for r in rows_due:
        vals.append(
            f"  ({sql_str(r['id'])}, {sql_str(r['account'])}, {sql_str(r['supplier'])}, "
            f"{sql_num(r['total'])}, {sql_num(r['paid'])}, {sql_num(r['due'])}, "
            f"{sql_str(SOURCE_FILE)}, {sql_str(r['source_sheet'])})"
        )
    out.append(',\n'.join(vals) + ';')
    out.append('')

# Due payments inserts
if rows_due_payment:
    out.append('-- ── mdr_account_due_payments ───────────────────────────────────')
    out.append('INSERT INTO public.mdr_account_due_payments')
    out.append('  (id, due_id, account, supplier, hijri_year, month, day, gregorian_date,')
    out.append('   amount, receipt_no, source_file, source_sheet, source_row)')
    out.append('VALUES')
    vals = []
    for r in rows_due_payment:
        vals.append(
            f"  ({sql_str(r['id'])}, {sql_str(r['due_id'])}, {sql_str(r['account'])}, "
            f"{sql_str(r['supplier'])}, {sql_str(r['hijri_year'])}, {sql_str(r['month'])}, "
            f"{sql_int(r['day'])}, {sql_date(r['gregorian_date'])}, "
            f"{sql_num(r['amount'])}, {sql_str(r['receipt_no'])}, "
            f"{sql_str(SOURCE_FILE)}, {sql_str(r['source_sheet'])}, {sql_int(r['source_row'])})"
        )
    out.append(',\n'.join(vals) + ';')
    out.append('')

out.append('-- ═══════════════════════════════════════════════════════════════')
out.append(f'-- Import complete: {len(rows_income)} incomes, {len(rows_expense)} expenses,')
out.append(f'-- {len(rows_due)} dues, {len(rows_due_payment)} due_payments')
out.append('-- ═══════════════════════════════════════════════════════════════')

sql_text = '\n'.join(out)
out_path = r'd:\programming\11-idarah\tools\import_accounts_1447_48.sql'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(sql_text)

# Print summary to stdout
print('=== SQL Generation Complete ===')
print(f'Income rows  : {len(rows_income)}')
print(f'Expense rows : {len(rows_expense)}')
expense_by_acct = {}
for r in rows_expense:
    expense_by_acct[r["account"]] = expense_by_acct.get(r["account"], 0) + 1
for acct, cnt in sorted(expense_by_acct.items()):
    print(f'  {acct:12s}: {cnt}')
print(f'Due rows     : {len(rows_due)}')
print(f'Due payments : {len(rows_due_payment)}')
print()

# Verify some totals
matbakh_total = sum(r['amount'] for r in rows_expense if r['account']=='matbakh')
madrasa_total = sum(r['amount'] for r in rows_expense if r['account']=='madrasa')
tamirat_total = sum(r['amount'] for r in rows_expense if r['account']=='tamirat')
general_total = sum(r['amount'] for r in rows_expense if r['account']=='general')
income_total  = sum(r['amount'] for r in rows_income)
print('=== আর্থিক যাচাই ===')
print(f'আয় মোট     : {income_total:,.0f}')
print(f'মাতবাখ ব্যয়: {matbakh_total:,.0f}  (Excel: 2,135,165)')
print(f'মাদরাসা ব্যয়: {madrasa_total:,.0f}  (Excel: 1,219,694)')
print(f'তামিরাত ব্যয়: {tamirat_total:,.0f}  (Excel: 872,108)')
print(f'সাধারণ ব্যয় : {general_total:,.0f}')
print(f'মোট ব্যয়    : {matbakh_total+madrasa_total+tamirat_total+general_total:,.0f}')
print()
print(f'SQL file: {out_path}')
