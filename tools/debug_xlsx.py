import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

path = r'C:\Users\sanim\Downloads\মাতবাখ-মাদরাসা-৪৭-৪৮.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

# Check মাতবাখ rows 6-30
print('=== মাতবাখ rows 6-30 (data_only) ===')
ws = wb['মাতবাখ']
for rno, row in enumerate(ws.iter_rows(min_row=6, max_row=30, values_only=True), start=6):
    print(f'row{rno}: A={repr(row[0])} B={repr(row[1])} C={repr(row[2])} D={repr(row[3])} E={repr(row[4])} F={repr(row[5])} G={repr(row[6])} H={repr(row[7])}')

print()

# Find first শাওয়াল and জিলকদ rows
print('=== First non-empty month transitions ===')
ws = wb['মাতবাখ']
prev_month = None
row_count = 0
shawwal_count = 0
jilkad_count = 0
for rno, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
    m = row[0]
    amt = row[6]
    if m and str(m).strip():
        month = str(m).strip()
        if month != prev_month:
            print(f'row{rno}: month changed to {repr(month)}, amt={repr(amt)}')
            prev_month = month
        if 'শাওয়াল' in month or 'Shawwal' in month.lower():
            if amt and isinstance(amt, (int, float)):
                shawwal_count += 1
        if 'জিলক' in month:
            if amt and isinstance(amt, (int, float)):
                jilkad_count += 1
    row_count += 1
    if row_count > 3000:
        break

print(f'শাওয়াল rows with amount: {shawwal_count}')
print(f'জিলকদ rows with amount: {jilkad_count}')
print()

# Check what num_val does with typical amounts
def num_val(v):
    if v is None:
        return None
    try:
        f = float(str(v).replace(',',''))
        return f if f != 0 else None
    except:
        return None

print('=== Amount column type check for first 20 data rows ===')
ws = wb['মাতবাখ']
count = 0
for rno, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
    amt = row[6]
    m = row[0]
    if m or amt:
        print(f'row{rno}: month={repr(m)} amt_raw={repr(amt)} type={type(amt).__name__} num_val={repr(num_val(amt))}')
        count += 1
        if count >= 20:
            break
