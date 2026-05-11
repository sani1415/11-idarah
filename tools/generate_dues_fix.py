"""
Dues fix script — regenerate only mdr_account_dues + mdr_account_due_payments
with corrected column mapping for the right-side (tamirat) suppliers.
"""
import openpyxl, sys, io, re, unicodedata
from datetime import date, datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

EXCEL_PATH = r'C:\Users\sanim\Downloads\মাতবাখ-মাদরাসা-৪৭-৪৮.xlsx'

# generate_accounts_import.py এর মতো — গ্রেগরিয়ান→হিজরী মাস + আনুমানিক দিন
GREG_HIJRI_RANGES = [
    (date(2025, 7, 7),  date(2025, 8, 5),  'মুহাররম', 1447),
    (date(2025, 8, 6),  date(2025, 9, 3),  'সফর', 1447),
    (date(2025, 9, 4),  date(2025, 10, 3), 'রবিউল আউয়াল', 1447),
    (date(2025, 10, 4), date(2025, 11, 1), 'রবিউস সানি', 1447),
    (date(2025, 11, 2), date(2025, 11, 30),'জুমাদাল উলা', 1447),
    (date(2025, 12, 1), date(2025, 12, 29),'জুমাদাল উখরা', 1447),
    (date(2025, 12, 30),date(2026, 1, 27), 'রজব', 1447),
    (date(2026, 1, 28), date(2026, 2, 25), 'শাবান', 1447),
    (date(2026, 2, 26), date(2026, 3, 28), 'রমজান', 1447),
    (date(2026, 3, 29), date(2026, 4, 26), 'শাওয়াল', 1447),
    (date(2026, 4, 27), date(2026, 5, 25), 'জিলকদ', 1447),
    (date(2026, 5, 26), date(2026, 6, 24), 'জিলহজ', 1447),
    (date(2026, 6, 25), date(2026, 7, 23), 'মুহাররম', 1448),
    (date(2026, 7, 24), date(2026, 8, 21), 'সফর', 1448),
]

def greg_to_hijri_full(g):
    if not g:
        return None, None, None
    if isinstance(g, datetime):
        g = g.date()
    if not isinstance(g, date):
        return None, None, None
    for start, end, month, year in GREG_HIJRI_RANGES:
        if start <= g <= end:
            day_h = min(30, max(1, (g - start).days + 1))
            return month, year, day_h
    return None, None, None

def nfc(s):
    return unicodedata.normalize('NFC', str(s)).replace('\u200c','').replace('\u200d','').strip()

def str_val(v):
    if v is None: return None
    s = nfc(str(v)).strip()
    return s if s else None

def num_val(v):
    if v is None: return None
    try:
        f = float(str(v).replace(',',''))
        return f if f != 0 else None
    except: return None

HIJRI_MONTHS = ['মুহাররম','সফর','রবিউল আউয়াল','রবিউস সানি',
                'জুমাদাল উলা','জুমাদাল উখরা','রজব','শাবান',
                'রমজান','শাওয়াল','জিলকদ','জিলহজ']

SUPPLIER_ACCOUNT = {
    'সাগর':'matbakh','আল মদিনা':'matbakh','গ্যাস':'matbakh',
    'ডিম':'matbakh','বেলাল':'matbakh','লাকড়ি':'matbakh',
    'জমিনের সবজী':'matbakh',
    'আনোয়ার':'tamirat','বিশাল থাই':'tamirat','কামাল সিমেন্ট':'tamirat',
    'বন্ধু ট্রেডার্স':'tamirat','ঊষা স্যানেটারি':'tamirat',
    'বেলাল আলীপুর':'tamirat','মা হার্ডওয়ার':'tamirat','কাঠ মিস্ত্রি':'tamirat',
}

# Sequential IDs matching original script's gen_id pattern
_counter = [900]
def gen_id(prefix):
    _counter[0] += 1
    return f"{prefix}{_counter[0]:05d}"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['বাকির হিসাব']

# Read rows 2-9 (summary section)
due_rows_raw = list(ws.iter_rows(min_row=2, max_row=9, values_only=True))

# Debug: print raw right-side columns
print("=== Raw right-side data (col I=row[8], J=row[9], K=row[10], L=row[11]) ===")
for i, row in enumerate(due_rows_raw):
    rpad = [row[j] if j < len(row) else None for j in range(8, 13)]
    print(f"  Row {i+2}: H={rpad[0]!r:15} I={rpad[1]!r:25} J={rpad[2]!r:15} K={rpad[3]!r:15} L={rpad[4]!r}")

rows_due = []
all_due_rows = []

for row in due_rows_raw:
    # Left side: col B=no(row[1]), C=supplier(row[2]), D=total(row[3]), E=paid(row[4]), F=due(row[5])
    sup_left  = str_val(row[2])
    tot_left  = num_val(row[3])
    paid_left = num_val(row[4])
    due_left  = num_val(row[5])
    if sup_left and tot_left is not None:
        all_due_rows.append((sup_left, tot_left, paid_left, due_left))

    # Right side: col H=no(row[7]), I=supplier(row[8]), J=total(row[9]), K=paid(row[10]), L=due(row[11])
    if len(row) >= 10:
        sup_right  = str_val(row[8])
        tot_right  = num_val(row[9])
        paid_right = num_val(row[10])
        due_right  = num_val(row[11]) if len(row) > 11 else None
        if sup_right and tot_right is not None:
            all_due_rows.append((sup_right, tot_right, paid_right, due_right))

print(f"\n=== Dues rows found: {len(all_due_rows)} ===")
for s, t, p, d in all_due_rows:
    print(f"  {s:25} total={t}  paid={p}  due={d}")

for sup, total, paid, due in all_due_rows:
    if not sup: continue
    acct    = SUPPLIER_ACCOUNT.get(sup, 'matbakh')
    total_v = total or 0
    paid_v  = paid or 0
    due_v   = due if due is not None else (total_v - paid_v)
    rows_due.append({
        'id':      gen_id('due-47-'),
        'account': acct,
        'supplier': sup,
        'total':   total_v,
        'paid':    paid_v,
        'due':     due_v,
    })

# Payment rows (row 12+)
rows_due_payment = []
for rno, row in enumerate(ws.iter_rows(min_row=12, values_only=True), start=12):
    greg_date = row[1]
    amount    = num_val(row[2])
    supplier  = str_val(row[3])
    receipt   = str_val(row[4])
    if not amount or not supplier: continue
    greg = None
    if isinstance(greg_date, (date, datetime)):
        greg = greg_date if isinstance(greg_date, date) else greg_date.date()
    month, hyear, hday = greg_to_hijri_full(greg)
    if not month:
        month, hyear, hday = 'শাওয়াল', 1447, 1
    acct = SUPPLIER_ACCOUNT.get(supplier, 'matbakh')
    due_id = None
    for d in rows_due:
        if d['supplier'] == supplier and d['account'] == acct:
            due_id = d['id']
            break
    rows_due_payment.append({
        'id':           gen_id('dpay-47-'),
        'due_id':       due_id,
        'account':      acct,
        'supplier':     supplier,
        'hijri_year':   str(hyear),
        'month':        month,
        'day':          hday,
        'gregorian_date': greg,
        'amount':       amount,
        'receipt_no':   receipt,
        'source_row':   rno,
    })

print(f"\n=== Payments found: {len(rows_due_payment)} ===")
for p in rows_due_payment:
    print(f"  {p['supplier']:25} {p['amount']:>10}  due_id={p['due_id']}  acct={p['account']}")

# Generate SQL
def q(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'","''") + "'"

out = []
out.append('-- dues fix: correct tamirat supplier mapping')
out.append('-- Step 1: remove wrong due_payments')
out.append('DELETE FROM public.mdr_account_due_payments;')
out.append('')
out.append('-- Step 2: remove all dues')
out.append('DELETE FROM public.mdr_account_dues;')
out.append('')

# INSERT dues
if rows_due:
    out.append('-- Step 3: insert corrected dues')
    cols = '(id, account, supplier, total, paid, due, source_sheet)'
    vals = []
    for r in rows_due:
        vals.append(f"({q(r['id'])},{q(r['account'])},{q(r['supplier'])},{r['total']},{r['paid']},{r['due']},{q('বাকির হিসাব')})")
    out.append(f"INSERT INTO public.mdr_account_dues {cols} VALUES")
    out.append(',\n'.join(vals) + ';')
    out.append('')

# INSERT due_payments
if rows_due_payment:
    out.append('-- Step 4: insert corrected due_payments')
    cols = '(id, due_id, account, supplier, hijri_year, month, day, gregorian_date, amount, receipt_no, source_sheet, source_row)'
    vals = []
    for r in rows_due_payment:
        gd = f"'{r['gregorian_date']}'" if r['gregorian_date'] else 'NULL'
        day_sql = str(int(r['day'])) if r.get('day') is not None else 'NULL'
        vals.append(
            f"({q(r['id'])},{q(r['due_id'])},{q(r['account'])},{q(r['supplier'])},"
            f"{q(r['hijri_year'])},{q(r['month'])},{day_sql},{gd},{r['amount']},"
            f"{q(r['receipt_no'])},{q('বাকির হিসাব')},{r['source_row']})"
        )
    out.append(f"INSERT INTO public.mdr_account_due_payments {cols} VALUES")
    out.append(',\n'.join(vals) + ';')
    out.append('')

out.append('-- Verify')
out.append("SELECT 'dues' as tbl, count(*) as rows, sum(total) as total_sum, sum(due) as due_sum FROM mdr_account_dues;")
out.append("SELECT 'due_payments' as tbl, count(*) as rows, sum(amount) FROM mdr_account_due_payments;")

sql_out = '\n'.join(out)
with open('tools/dues_fix.sql', 'w', encoding='utf-8') as f:
    f.write(sql_out)
print(f"\n\nSQL written to tools/dues_fix.sql ({len(rows_due)} dues, {len(rows_due_payment)} payments)")
