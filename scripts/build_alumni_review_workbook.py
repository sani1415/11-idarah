"""
build_alumni_review_workbook.py
================================
Input : outputs/alumni_merged/alumni_clean_combined.xlsx
Output: outputs/alumni_merged/alumni_review_workbook_v1.xlsx

দুটো শিট তৈরি করে:
  1. "স্বয়ংক্রিয়_মার্জ"  → কোনো hard-conflict নেই, auto-merge করা হয়েছে
  2. "রিভিউ_দরকার"        → hard-conflict আছে, প্রতিটি গ্রুপে:
        • সবুজ "প্রস্তাবিত মার্জ" রো (সম্পাদনযোগ্য)
        • ধূসর source রো (শুধু দেখার জন্য)
        • conflict আছে এমন সেলগুলো হলুদ/কমলা রঙে চিহ্নিত

ব্যবহার:
  python scripts/build_alumni_review_workbook.py
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = 'outputs/alumni_merged/alumni_clean_combined.xlsx'
OUTPUT_FILE = 'outputs/alumni_merged/alumni_review_workbook_v1.xlsx'

# ── ডুপ্লিকেট শিটের কলাম সূচক (0-ভিত্তিক) ──────────────────────────────
CI_GROUP   = 0   # ডুপ্লিকেট গ্রুপ
CI_REASON  = 1   # মিলের কারণ
CI_SRCROW  = 2   # মূল শীট Row
CI_JAMAT   = 3   # জামাত নং
CI_ID      = 4   # পার্মানেন্ট আইডি
CI_NAME    = 5   # নাম
CI_FATHER  = 6   # বাবার নাম
CI_DIST    = 7   # জেলা
CI_THANA   = 8   # উপজেলা/থানা
CI_ADDR    = 9   # বিস্তারিত ঠিকানা
CI_STATUS  = 10  # বর্তমান অবস্থা/হালাত
CI_MOB1    = 11  # মোবাইল ১
CI_MOB2    = 12  # মোবাইল ২
CI_MOB3    = 13  # অতিরিক্ত মোবাইল
CI_STYEAR  = 14  # অধ্যয়ন শুরু বর্ষ
CI_ENDYR   = 15  # অধ্যয়ন শেষ বর্ষ
CI_STDESC  = 16  # অধ্যয়ন বিবরণ
CI_GSTAT   = 17  # স্ট্যাটাস
CI_NOTES   = 18  # মন্তব্য
CI_SRC     = 19  # ডেটা উৎস

# মানুষের সিদ্ধান্ত দরকার এমন কলাম
HARD_COLS = [CI_JAMAT, CI_ID, CI_NAME, CI_FATHER, CI_DIST, CI_THANA,
             CI_STATUS, CI_STYEAR, CI_ENDYR, CI_GSTAT]

# স্বয়ংক্রিয়ভাবে সমাধানযোগ্য কলাম
SOFT_LONGEST = [CI_ADDR, CI_STDESC]   # লম্বাটা নেওয়া হবে
SOFT_PHONES  = [CI_MOB1, CI_MOB2, CI_MOB3]  # unique নম্বর সংগ্রহ
SOFT_CONCAT  = [CI_NOTES]             # জোড়া দেওয়া হবে " | " দিয়ে

# Clean List শিটের হেডার অর্ডার (3..19 → 0..16)
DATA_HEADERS = [
    'জামাত নং', 'পার্মানেন্ট আইডি', 'নাম', 'বাবার নাম',
    'জেলা', 'উপজেলা/থানা', 'বিস্তারিত ঠিকানা', 'বর্তমান অবস্থা/হালাত',
    'মোবাইল ১', 'মোবাইল ২', 'অতিরিক্ত মোবাইল',
    'অধ্যয়ন শুরু বর্ষ', 'অধ্যয়ন শেষ বর্ষ', 'অধ্যয়ন বিবরণ',
    'স্ট্যাটাস', 'মন্তব্য', 'ডেটা উৎস',
]

# ── রঙ ────────────────────────────────────────────────────────────────────
FILL_MERGE_ROW  = PatternFill('solid', fgColor='C8F7C5')   # হালকা সবুজ
FILL_SRC_ROW    = PatternFill('solid', fgColor='F0F0F0')   # হালকা ধূসর
FILL_CONFLICT   = PatternFill('solid', fgColor='FFD966')   # হলুদ
FILL_HARD_CONF  = PatternFill('solid', fgColor='FF9900')   # কমলা (hard)
FILL_GROUP_HDR  = PatternFill('solid', fgColor='4472C4')   # নীল
FILL_AUTO_HDR   = PatternFill('solid', fgColor='375623')   # গাঢ় সবুজ

FONT_WHITE_BOLD = Font(bold=True, color='FFFFFF', size=10)
FONT_MERGE      = Font(bold=True, size=10)
FONT_SRC        = Font(italic=True, size=9, color='555555')
FONT_HEADER     = Font(bold=True, size=10)


# ── ইউটিলিটি ──────────────────────────────────────────────────────────────

def nz(v):
    """None/empty → ''"""
    if v is None:
        return ''
    return str(v).strip()

def unique_vals(rows, col):
    seen, out = set(), []
    for r in rows:
        v = nz(r[col])
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out

def has_conflict(rows, col):
    return len(unique_vals(rows, col)) > 1

def merge_phones(rows):
    """তিনটি ফোন স্লটে unique নম্বর ভরে দাও।"""
    phones, seen = [], set()
    for r in rows:
        for c in SOFT_PHONES:
            v = nz(r[c])
            if v and v not in seen:
                seen.add(v)
                phones.append(v)
    result = [None, None, None]
    for i, p in enumerate(phones[:3]):
        result[i] = p
    return result

def build_proposed(rows):
    """
    প্রস্তাবিত মার্জ রো তৈরি কর।
    Hard conflict-এ: প্রথম non-empty মান রাখো (রিভিউ দরকার)।
    Soft-এ: auto-resolve।
    Returns: list[20] (ডুপ্লিকেট শিটের column layout)
    """
    merged = [None] * 20
    # সব কলামে প্রথম non-empty মান দিয়ে শুরু
    for col in range(20):
        vals = unique_vals(rows, col)
        if vals:
            merged[col] = vals[0]
    # Soft: ফোন
    ph = merge_phones(rows)
    merged[CI_MOB1], merged[CI_MOB2], merged[CI_MOB3] = ph
    # Soft: লম্বাটা
    for col in SOFT_LONGEST:
        vals = [nz(r[col]) for r in rows if nz(r[col])]
        if vals:
            merged[col] = max(vals, key=len)
    # Soft: জোড়া
    for col in SOFT_CONCAT:
        parts = unique_vals(rows, col)
        merged[col] = ' | '.join(parts) if parts else None
    # ডেটা উৎস জোড়া
    srcs = unique_vals(rows, CI_SRC)
    merged[CI_SRC] = ' + '.join(srcs) if srcs else None
    return merged

def can_auto_merge(rows):
    """Hard conflict নেই → True"""
    return not any(has_conflict(rows, col) for col in HARD_COLS)

def do_auto_merge(rows):
    """নিরাপদ auto-merge।"""
    return build_proposed(rows)

def data_cols(row):
    """ডুপ্লিকেট শিটের রো থেকে 17টি data col বের করো।"""
    return list(row[CI_JAMAT: CI_SRC + 1])  # col 3..19


# ── শিট লেখার হেলপার ──────────────────────────────────────────────────────

def write_row(ws, row_idx, values, fill=None, font=None, wrap=False):
    for ci, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze_top(ws):
    ws.freeze_panes = 'A2'


# ── মূল কাজ ───────────────────────────────────────────────────────────────

def main():
    print('ফাইল পড়ছি...')
    wb_in = openpyxl.load_workbook(INPUT_FILE)
    ws_dup = wb_in['সম্ভাব্য ডুপ্লিকেট']

    # সব রো পড়ো
    all_rows = list(ws_dup.iter_rows(min_row=2, values_only=True))
    print(f'  মোট ডুপ্লিকেট রো: {len(all_rows)}')

    # গ্রুপে ভাগ করো
    groups = {}
    for r in all_rows:
        g = r[CI_GROUP]
        groups.setdefault(g, []).append(r)
    print(f'  মোট গ্রুপ: {len(groups)}')

    auto_groups   = {}
    review_groups = {}
    for gid, rows in groups.items():
        if can_auto_merge(rows):
            auto_groups[gid]   = rows
        else:
            review_groups[gid] = rows

    print(f'  স্বয়ংক্রিয় মার্জযোগ্য: {len(auto_groups)} গ্রুপ')
    print(f'  রিভিউ দরকার         : {len(review_groups)} গ্রুপ')

    wb_out = openpyxl.Workbook()

    # ── Sheet 1: স্বয়ংক্রিয়_মার্জ ──────────────────────────────────────
    ws_auto = wb_out.active
    ws_auto.title = 'স্বয়ংক্রিয়_মার্জ'

    hdr = ['গ্রুপ', 'মিলের কারণ', 'উৎস রো'] + DATA_HEADERS
    write_row(ws_auto, 1, hdr, fill=FILL_AUTO_HDR, font=FONT_WHITE_BOLD)
    freeze_top(ws_auto)

    row_i = 2
    for gid, rows in auto_groups.items():
        merged = do_auto_merge(rows)
        reason = nz(rows[0][CI_REASON])
        src_rows = ', '.join(str(r[CI_SRCROW]) for r in rows)
        out = [gid, reason, src_rows] + data_cols(merged)
        write_row(ws_auto, row_i, out, fill=FILL_MERGE_ROW, font=FONT_MERGE, wrap=True)
        row_i += 1

    set_col_widths(ws_auto, [8, 20, 12] + [14, 12, 18, 18, 12, 14, 28,
                                             36, 14, 14, 14, 10, 10, 30, 12, 30, 14])

    # ── Sheet 2: রিভিউ_দরকার ─────────────────────────────────────────────
    ws_rev = wb_out.create_sheet('রিভিউ_দরকার')

    # হেডার রো
    rev_hdr = ['ধরন', 'গ্রুপ', 'conflict/উৎস'] + DATA_HEADERS
    write_row(ws_rev, 1, rev_hdr, fill=FILL_GROUP_HDR, font=FONT_WHITE_BOLD)
    freeze_top(ws_rev)

    row_i = 2
    for gid, rows in review_groups.items():
        # কোন কলামে conflict আছে চিহ্নিত করো
        conflicted_data_cols = set()
        for col in HARD_COLS:
            if has_conflict(rows, col):
                conflicted_data_cols.add(col - CI_JAMAT)  # 0-based data index

        conflict_names = [DATA_HEADERS[c] for c in conflicted_data_cols]
        reason_str = nz(rows[0][CI_REASON])
        conflict_str = ', '.join(conflict_names)

        # গ্রুপ হেডার রো
        proposed = build_proposed(rows)
        grp_label = f'গ্রুপ {gid} | {reason_str} | conflict: {conflict_str}'
        ws_rev.merge_cells(
            start_row=row_i, start_column=1,
            end_row=row_i, end_column=len(rev_hdr)
        )
        hdr_cell = ws_rev.cell(row=row_i, column=1, value=grp_label)
        hdr_cell.fill = FILL_GROUP_HDR
        hdr_cell.font = FONT_WHITE_BOLD
        row_i += 1

        # প্রস্তাবিত মার্জ রো (সবুজ)
        merge_out = ['✏️ মার্জ', gid, conflict_str] + data_cols(proposed)
        for ci, val in enumerate(merge_out, start=1):
            cell = ws_rev.cell(row=row_i, column=ci, value=val)
            cell.fill = FILL_MERGE_ROW
            cell.font = FONT_MERGE
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # conflict আছে এমন data col কমলা করো
            data_ci = ci - 4  # shift: col 4 = data index 0
            if 0 <= data_ci < len(DATA_HEADERS) and data_ci in conflicted_data_cols:
                cell.fill = FILL_HARD_CONF
        row_i += 1

        # source রো (ধূসর)
        for r in rows:
            src_label = nz(r[CI_SRC]) or nz(r[CI_REASON])
            src_out = ['📋 উৎস', gid, src_label] + data_cols(r)
            for ci, val in enumerate(src_out, start=1):
                cell = ws_rev.cell(row=row_i, column=ci, value=val)
                cell.fill = FILL_SRC_ROW
                cell.font = FONT_SRC
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # conflict column হলুদ করো
                data_ci = ci - 4
                if 0 <= data_ci < len(DATA_HEADERS) and data_ci in conflicted_data_cols:
                    cell.fill = FILL_CONFLICT
            row_i += 1

        # ফাঁকা separator রো
        row_i += 1

    set_col_widths(ws_rev, [10, 8, 25] + [14, 12, 18, 18, 12, 14, 28,
                                            36, 14, 14, 14, 10, 10, 30, 12, 30, 14])

    # ── সংরক্ষণ ───────────────────────────────────────────────────────────
    print(f'\nসংরক্ষণ করছি → {OUTPUT_FILE}')
    wb_out.save(OUTPUT_FILE)
    print('সম্পন্ন!')
    print(f'\nসারসংক্ষেপ:')
    print(f'  স্বয়ংক্রিয়_মার্জ শিট  : {len(auto_groups)} রো')
    print(f'  রিভিউ_দরকার শিট       : {len(review_groups)} গ্রুপ')
    print(f'  রিভিউ শিটে মোট রো     : ~{sum(len(v)+3 for v in review_groups.values())}')


if __name__ == '__main__':
    main()
