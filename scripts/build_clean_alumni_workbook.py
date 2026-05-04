from __future__ import annotations

import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OLD_FILE = ROOT / "alumni.xlsx"
NEW_FILE = ROOT / "new_alumni.xlsx"
OUT_DIR = ROOT / "outputs" / "alumni_merged"
OUT_FILE = OUT_DIR / "alumni_clean_combined_v5.xlsx"

BN_DIGITS = "০১২৩৪৫৬৭৮৯"
EN_DIGITS = "0123456789"
TRANS_DIGITS = str.maketrans(BN_DIGITS, EN_DIGITS)

BIN = "বিন"
IBNE = "ইবনে"
PUTRA = "পুত্র"

HEADERS = [
    "জামাত নং",
    "পার্মানেন্ট আইডি",
    "নাম",
    "বাবার নাম",
    "জেলা",
    "উপজেলা/থানা",
    "বিস্তারিত ঠিকানা",
    "বর্তমান অবস্থা/হালাত",
    "মোবাইল ১",
    "মোবাইল ২",
    "অতিরিক্ত মোবাইল",
    "অধ্যয়ন শুরু বর্ষ",
    "অধ্যয়ন শেষ বর্ষ",
    "অধ্যয়ন বিবরণ",
    "স্ট্যাটাস",
    "মন্তব্য",
    "ডেটা উৎস",
]

DUPLICATE_HEADERS = [
    "ডুপ্লিকেট গ্রুপ",
    "মিলের কারণ",
    "মূল শীট Row",
    "জামাত নং",
    "পার্মানেন্ট আইডি",
    "নাম",
    "বাবার নাম",
    "মোবাইল ১",
    "মোবাইল ২",
    "অধ্যয়ন বিবরণ",
    "মন্তব্য",
    "ডেটা উৎস",
]

MERGED_MATCH_HEADERS = [
    "ম্যাচ নং",
    "মিলের ধরন",
    "মূল শীট Row",
    "নতুন তালিকা Row",
    "পুরনো তালিকা Row",
    "জামাত নং",
    "পার্মানেন্ট আইডি",
    "নতুন নাম",
    "পুরনো নাম",
    "বাবার নাম",
    "নতুন মোবাইল",
    "পুরনো মোবাইল",
    "অধ্যয়ন বিবরণ",
    "পুরনো মন্তব্য",
    "মন্তব্য",
    "ডেটা উৎস",
]

HONORIFIC_TOKENS = {
    "মুহাম্মদ", "মুহাম্মাদ", "মোহাম্মদ", "মোঃ", "মো", "হাফেজ", "মাওলানা",
    "মাওঃ", "মাও", "আলহাজ", "আলহাজ্ব", "মুফতী", "মুফতি", "মরহুম",
    "হাজী", "হাজি", "রহ", "রহঃ", "রাহঃ", "সাহেব", "আল",
}

DISTRICTS = [
    "ঢাকা", "কুমিল্লা", "মুন্সীগঞ্জ", "নোয়াখালী", "নোয়াখালী", "লক্ষ্মীপুর", "ময়মনসিংহ",
    "মোমেনশাহী", "গোপালগঞ্জ", "চাঁদপুর", "বরিশাল", "ভোলা", "ফেনী", "চট্টগ্রাম",
    "নারায়ণগঞ্জ", "নারায়ণগঞ্জ", "গাজীপুর", "মানিকগঞ্জ", "ফরিদপুর", "মাদারীপুর",
    "শরীয়তপুর", "শরিয়তপুর", "কিশোরগঞ্জ", "নেত্রকোনা", "টাঙ্গাইল", "জামালপুর",
    "শেরপুর", "সিলেট", "সুনামগঞ্জ", "হবিগঞ্জ", "মৌলভীবাজার", "রাজশাহী", "বগুড়া",
    "বগুড়া", "নওগাঁ", "নাটোর", "চাঁপাইনবাবগঞ্জ", "রংপুর", "দিনাজপুর", "কুড়িগ্রাম",
    "কুড়িগ্রাম", "পঞ্চগড়", "পঞ্চগড়", "নীলফামারী", "খুলনা", "যশোর", "কুষ্টিয়া",
    "কুষ্টিয়া", "ঝিনাইদহ", "সাতক্ষীরা", "বাগেরহাট",
]

THANAS = [
    "কামরাঙ্গীরচর", "কামরাঙ্গীর চর", "মিরপুর", "হাজারীবাগ", "উত্তরা", "ডেমরা",
    "বাড্ডা", "গুলশান", "মোহাম্মদপুর", "কেরানীগঞ্জ", "দক্ষিণ কেরানীগঞ্জ",
    "সাভার", "আশুলিয়া", "আশুলিয়া", "তুরাগ", "খিলগাঁও", "রমনা", "রামপুরা",
    "শ্রীপুর", "গজারিয়া", "গজারিয়া", "চাটখিল", "চৌদ্দগ্রাম", "মুক্তাগাছা",
    "ঝিগাতলা", "আজিমপুর", "বনশ্রী", "সারুলিয়া",
]


def clean_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_text(value) -> str:
    text = unicodedata.normalize("NFKC", clean_value(value)).translate(TRANS_DIGITS)
    text = re.sub(r"[\u200c\u200d\ufeff]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def to_text_number(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def normalize_phone(value) -> list[str]:
    text = normalize_text(value)
    phones: list[str] = []
    for match in re.findall(r"\d{7,15}", text):
        phone = match
        if len(phone) == 10 and phone.startswith("1"):
            phone = "0" + phone
        if phone not in phones:
            phones.append(phone)
    return phones


def split_name_and_father(full_name: str) -> tuple[str, str]:
    name = clean_value(full_name)
    for marker in (f" {BIN} ", f" {IBNE} ", f" {PUTRA} "):
        if marker in name:
            left, right = name.split(marker, 1)
            return left.strip(" ,।"), right.strip(" ,।")
    return name.strip(" ,।"), ""


def base_name(full_name: str) -> str:
    name, _ = split_name_and_father(normalize_text(full_name))
    for ch in "।,;:()[]{}#-–—/\\._":
        name = name.replace(ch, " ")
    return re.sub(r"\s+", " ", name).strip()


def raw_tokens(full_name: str) -> list[str]:
    return [t for t in base_name(full_name).split() if len(t) > 1]


def duplicate_jamat_key(jamat: str) -> str:
    text = normalize_text(jamat)
    match = re.search(r"\d+", text)
    if match:
        return match.group(0)
    return text


def duplicate_name_parts(name: str) -> tuple[str, set[str]]:
    text = base_name(name)
    for ch in "।,;:()[]{}#-–—/\\._":
        text = text.replace(ch, " ")
    words = []
    for word in re.sub(r"\s+", " ", text).strip().split():
        if word in HONORIFIC_TOKENS:
            continue
        words.append(word)
    normalized = " ".join(words)
    return normalized, set(words)


def duplicate_pair_reason(a: dict, b: dict) -> str:
    if duplicate_jamat_key(a["jamat"]) != duplicate_jamat_key(b["jamat"]):
        return ""
    name_a, tokens_a = duplicate_name_parts(a["name"])
    name_b, tokens_b = duplicate_name_parts(b["name"])
    if not name_a or not name_b:
        return ""
    if name_a == name_b:
        return "একই জামাত + একই নাম"
    shorter, longer = sorted([name_a, name_b], key=len)
    if len(shorter) >= 5 and shorter in longer:
        return "একই জামাত + নামের অংশ মিলে"
    shared = tokens_a & tokens_b
    if len(shared) >= 2:
        score = len(shared) / max(len(tokens_a), len(tokens_b))
        if score >= 0.67:
            return "একই জামাত + নামের শব্দ মিলে"
    return ""


def find_duplicate_groups(records: list[dict]) -> list[tuple[int, str, list[dict]]]:
    parent = list(range(len(records)))
    reasons: dict[tuple[int, int], str] = {}

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int, reason: str) -> None:
        root_a = find(a)
        root_b = find(b)
        if root_a == root_b:
            return
        parent[root_b] = root_a
        reasons[tuple(sorted((a, b)))] = reason

    by_jamat: dict[str, list[int]] = defaultdict(list)
    for idx, record in enumerate(records):
        by_jamat[duplicate_jamat_key(record["jamat"])].append(idx)

    for indexes in by_jamat.values():
        for pos, left_idx in enumerate(indexes):
            for right_idx in indexes[pos + 1:]:
                reason = duplicate_pair_reason(records[left_idx], records[right_idx])
                if reason:
                    union(left_idx, right_idx, reason)

    grouped: dict[int, list[int]] = defaultdict(list)
    for idx in range(len(records)):
        grouped[find(idx)].append(idx)

    duplicate_groups = []
    group_no = 1
    for indexes in grouped.values():
        if len(indexes) < 2:
            continue
        group_reasons = set()
        for pos, left_idx in enumerate(indexes):
            for right_idx in indexes[pos + 1:]:
                reason = duplicate_pair_reason(records[left_idx], records[right_idx])
                if reason:
                    group_reasons.add(reason)
        duplicate_groups.append((group_no, " | ".join(sorted(group_reasons)), [records[i] for i in indexes]))
        group_no += 1
    return duplicate_groups


def detect_location(address: str) -> tuple[str, str, str, list[str]]:
    text = clean_value(address)
    district = ""
    thana = ""
    comments: list[str] = []
    for d in DISTRICTS:
        if d and d in text:
            district = "নোয়াখালী" if d == "নোয়াখালী" else d
            district = "নারায়ণগঞ্জ" if d == "নারায়ণগঞ্জ" else district
            break
    for t in THANAS:
        if t and t in text:
            thana = t
            break
    if text and not district:
        comments.append("জেলা যাচাই দরকার")
    if text and not thana:
        comments.append("উপজেলা/থানা যাচাই দরকার")
    return district, thana, text, comments


def parse_study_detail(detail: str) -> tuple[str, str]:
    text = clean_value(detail).replace("\xa0", " ")
    ordinals = re.findall(r"[১-৯]\s*(?:ম|য়|য়|র্থ|ষ্ঠ)?", text)
    ordinals = [re.sub(r"\s+", "", x) for x in ordinals]
    if not ordinals:
        if "পাঁচ বছর" in text or "পূর্ণ পাঁচ" in text:
            return "১ম", "৫ম"
        return "", ""
    if len(ordinals) == 1:
        return ordinals[0], ordinals[0]
    return ordinals[0], ordinals[-1]


def extract_jamat(batch: str) -> str:
    text = clean_value(batch)
    match = re.search(r"([০-৯]+(?:ম|য়|য়|র্থ|ষ্ঠ)?\s+জামাত)", text)
    if match:
        return match.group(1)
    return text


def parse_old_records() -> list[dict]:
    wb = load_workbook(OLD_FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records: list[dict] = []
    current = None
    for excel_row, row in enumerate(ws.iter_rows(min_row=9, max_col=9, values_only=True), start=9):
        values = list(row) + [""] * 9
        batch, serial, name, address, phone, halat, note, extra_1, extra_2 = values[:9]
        if isinstance(serial, (int, float)) and name:
            phones = normalize_phone(phone)
            student_name, father_name = split_name_and_father(clean_value(name))
            district, thana, detail_address, loc_comments = detect_location(clean_value(address))
            start_year, end_year = parse_study_detail(clean_value(note))
            current = {
                "source_row": excel_row,
                "jamat": extract_jamat(clean_value(batch)),
                "permanent_id": "",
                "name": student_name,
                "father_name": father_name,
                "district": district,
                "thana": thana,
                "address": detail_address,
                "halat": clean_value(halat),
                "phones": phones,
                "study_start": start_year,
                "study_end": end_year,
                "study_detail": clean_value(note),
                "status": "",
                "comments": loc_comments[:],
                "source": "পুরনো তালিকা",
                "raw_name": clean_value(name),
            }
            current["comments"].append("পার্মানেন্ট আইডি খুঁজে বসাতে হবে")
            records.append(current)
        elif current:
            continuation = [clean_value(v) for v in values[1:9] if v not in (None, "")]
            if continuation:
                current["halat"] = (current["halat"] + " " + " ".join(continuation)).strip()
    return records


def parse_new_records() -> list[dict]:
    wb = load_workbook(NEW_FILE, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    records: list[dict] = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, max_col=13, values_only=True), start=2):
        values = list(row) + [""] * 13
        seq, jamat, permanent_id, full_name, address, halat, phone_1, phone_2, study = values[:9]
        if not permanent_id and not full_name:
            continue
        phones = []
        for phone in normalize_phone(phone_1) + normalize_phone(phone_2):
            if phone not in phones:
                phones.append(phone)
        student_name, father_name = split_name_and_father(clean_value(full_name))
        district, thana, detail_address, loc_comments = detect_location(clean_value(address))
        start_year, end_year = parse_study_detail(clean_value(study))
        comments = loc_comments[:]
        if not study:
            comments.append("অধ্যয়ন বিবরণ খালি")
        if not halat:
            comments.append("বর্তমান অবস্থা/হালাত খালি")
        if not phones:
            comments.append("যোগাযোগ নম্বর খালি")
        records.append({
            "source_row": excel_row,
            "jamat": clean_value(jamat),
            "permanent_id": to_text_number(permanent_id),
            "name": student_name,
            "father_name": father_name,
            "district": district,
            "thana": thana,
            "address": detail_address,
            "halat": clean_value(halat),
            "phones": phones,
            "study_start": start_year,
            "study_end": end_year,
            "study_detail": clean_value(study),
            "status": "",
            "comments": comments,
            "source": "নতুন তালিকা",
            "raw_name": clean_value(full_name),
        })
    return records


def build_token_matcher(old_records: list[dict], new_records: list[dict]):
    all_tokens = []
    for record in old_records + new_records:
        all_tokens.extend(raw_tokens(record["raw_name"]))
    frequencies = Counter(all_tokens)
    common_tokens = {token for token, count in frequencies.items() if count >= 25}

    def record_tokens(record: dict) -> set[str]:
        return {token for token in raw_tokens(record["raw_name"]) if token not in common_tokens}

    for record in old_records + new_records:
        record["_tokens"] = record_tokens(record)


def add_unique_comment(record: dict, comment: str) -> None:
    if comment and comment not in record["comments"]:
        record["comments"].append(comment)


def phone_text(record: dict) -> str:
    return "; ".join(record.get("phones") or [])


def assign_review_groups(records: list[dict]) -> int:
    parent = list(range(len(records)))
    edge_reasons: dict[tuple[int, int], set[str]] = defaultdict(set)
    phone_edges: set[tuple[int, int]] = set()

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int, reason: str, is_phone: bool = False) -> None:
        if a == b:
            return
        pair = tuple(sorted((a, b)))
        edge_reasons[pair].add(reason)
        if is_phone:
            phone_edges.add(pair)
        root_a = find(a)
        root_b = find(b)
        if root_a != root_b:
            parent[root_b] = root_a

    new_indexes = [i for i, record in enumerate(records) if record.get("_source_kind") == "new"]
    old_indexes = [i for i, record in enumerate(records) if record.get("_source_kind") == "old"]

    old_by_phone: dict[str, list[int]] = defaultdict(list)
    for idx in old_indexes:
        for phone in records[idx]["phones"]:
            old_by_phone[phone].append(idx)
    for new_idx in new_indexes:
        for phone in records[new_idx]["phones"]:
            for old_idx in old_by_phone.get(phone, []):
                union(new_idx, old_idx, f"ফোন মিলেছে: {phone}", True)

    by_jamat: dict[str, list[int]] = defaultdict(list)
    for idx, record in enumerate(records):
        by_jamat[duplicate_jamat_key(record["jamat"])].append(idx)
    for indexes in by_jamat.values():
        for pos, left_idx in enumerate(indexes):
            for right_idx in indexes[pos + 1:]:
                reason = duplicate_pair_reason(records[left_idx], records[right_idx])
                if reason:
                    union(left_idx, right_idx, reason)

    by_permanent_id: dict[str, list[int]] = defaultdict(list)
    for idx, record in enumerate(records):
        if record["permanent_id"]:
            by_permanent_id[record["permanent_id"]].append(idx)
    for permanent_id, indexes in by_permanent_id.items():
        if len(indexes) < 2:
            continue
        for pos, left_idx in enumerate(indexes):
            for right_idx in indexes[pos + 1:]:
                union(left_idx, right_idx, f"পার্মানেন্ট আইডি duplicate: {permanent_id}")

    grouped: dict[int, list[int]] = defaultdict(list)
    for idx in range(len(records)):
        grouped[find(idx)].append(idx)

    review_groups = []
    for indexes in grouped.values():
        if len(indexes) < 2:
            continue
        group_edge_reasons = set()
        has_phone = False
        for pos, left_idx in enumerate(indexes):
            for right_idx in indexes[pos + 1:]:
                pair = tuple(sorted((left_idx, right_idx)))
                group_edge_reasons.update(edge_reasons.get(pair, set()))
                if pair in phone_edges:
                    has_phone = True
        if not group_edge_reasons:
            continue
        review_groups.append((min(records[i]["_original_sort"] for i in indexes), has_phone, sorted(group_edge_reasons), indexes))

    review_groups.sort(key=lambda item: item[0])
    for group_no, (_, has_phone, reasons, indexes) in enumerate(review_groups, start=1):
        color = "red" if has_phone else "yellow"
        reason_text = " | ".join(reasons)
        for idx in indexes:
            records[idx]["_review_group_no"] = group_no
            records[idx]["_review_color"] = color
            records[idx]["_review_reason"] = reason_text
    return len(review_groups)


def sort_records_for_review(records: list[dict]) -> list[dict]:
    def sort_key(record: dict):
        group_no = record.get("_review_group_no")
        source_rank = 0 if record.get("_source_kind") == "new" else 1
        if group_no:
            return (0, group_no, source_rank, record["_original_sort"])
        return (1, record["_original_sort"])

    return sorted(records, key=sort_key)


def add_review_comments(records: list[dict]) -> None:
    groups: dict[int, list[dict]] = defaultdict(list)
    for record in records:
        if record.get("_review_group_no"):
            groups[record["_review_group_no"]].append(record)

    for group_no, group_records in groups.items():
        for record in group_records:
            others = [other for other in group_records if other is not record]
            other_text = "; ".join(
                f"row {other['_clean_row']}: {other['name']}" + (f" ({phone_text(other)})" if phone_text(other) else "")
                for other in others[:6]
            )
            extra = "" if len(others) <= 6 else f"; আরও {len(others) - 6}টি"
            label = "লাল" if record.get("_review_color") == "red" else "হলুদ"
            add_unique_comment(
                record,
                f"{label}: Review group {group_no}; কারণ: {record.get('_review_reason', '')}; মিলেছে {other_text}{extra}",
            )


def combine_records(old_records: list[dict], new_records: list[dict]) -> list[dict]:
    build_token_matcher(old_records, new_records)

    output: list[dict] = []
    for idx, record in enumerate(new_records):
        row = dict(record)
        row["_source_kind"] = "new"
        row["_original_sort"] = idx
        output.append(row)
    for idx, record in enumerate(old_records):
        row = dict(record)
        row["_source_kind"] = "old"
        row["_original_sort"] = len(new_records) + idx
        output.append(row)

    assign_review_groups(output)
    output = sort_records_for_review(output)
    for clean_row, record in enumerate(output, start=2):
        record["_clean_row"] = clean_row
    add_review_comments(output)
    return output


def flatten_row(record: dict) -> list[str]:
    phones = record["phones"]
    return [
        record["jamat"],
        record["permanent_id"],
        record["name"],
        record["father_name"],
        record["district"],
        record["thana"],
        record["address"],
        record["halat"],
        phones[0] if len(phones) > 0 else "",
        phones[1] if len(phones) > 1 else "",
        "; ".join(phones[2:]) if len(phones) > 2 else "",
        record["study_start"],
        record["study_end"],
        record["study_detail"],
        record["status"],
        " | ".join(dict.fromkeys([c for c in record["comments"] if c])),
        record["source"],
    ]


def style_header(row, fill, font, border) -> None:
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def style_body(ws, border) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Nirmala UI", size=10)


def apply_review_fills(ws, records: list[dict]) -> tuple[int, int]:
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_count = 0
    yellow_count = 0
    for record in records:
        color = record.get("_review_color")
        if color not in ("red", "yellow"):
            continue
        fill = red_fill if color == "red" else yellow_fill
        if color == "red":
            red_count += 1
        else:
            yellow_count += 1
        row_idx = record.get("_clean_row")
        if not row_idx:
            continue
        for cell in ws[row_idx]:
            cell.fill = fill
    return red_count, yellow_count


def write_duplicate_sheet(wb: Workbook, records: list[dict], header_fill, header_font, border) -> int:
    duplicate_groups: list[tuple[int, str, list[dict]]] = []
    by_group: dict[int, list[dict]] = defaultdict(list)
    for record in records:
        if record.get("_review_group_no"):
            by_group[record["_review_group_no"]].append(record)
    for group_no in sorted(by_group):
        group_records = by_group[group_no]
        reason = group_records[0].get("_review_reason", "")
        duplicate_groups.append((group_no, reason, group_records))

    ws = wb.create_sheet("সম্ভাব্য ডুপ্লিকেট")
    ws.append(DUPLICATE_HEADERS)
    for group_no, reason, group_records in duplicate_groups:
        for record in sorted(group_records, key=lambda r: r.get("_clean_row", 0)):
            phones = record["phones"]
            ws.append([
                group_no,
                reason,
                record.get("_clean_row", ""),
                record["jamat"],
                record["permanent_id"],
                record["name"],
                record["father_name"],
                phones[0] if len(phones) > 0 else "",
                phones[1] if len(phones) > 1 else "",
                record["study_detail"],
                " | ".join(dict.fromkeys([c for c in record["comments"] if c])),
                record["source"],
            ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(ws.max_row, 1)}"
    style_header(ws[1], header_fill, header_font, border)
    style_body(ws, border)
    widths = {
        "A": 16, "B": 28, "C": 12, "D": 14, "E": 16, "F": 28,
        "G": 30, "H": 15, "I": 15, "J": 26, "K": 50, "L": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 42
    if ws.max_row >= 2:
        table = Table(displayName="PossibleDuplicateTable", ref=f"A1:L{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium7",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws.sheet_view.showGridLines = False
    return len(duplicate_groups)


def write_merged_match_sheet(wb: Workbook, records: list[dict], header_fill, header_font, border) -> int:
    matched_records = [
        record for record in records
        if record.get("_matched_old_row")
        or record["source"] in ("পুরনো+নতুন মিলিত", "সম্ভাব্য মিল, যাচাই দরকার")
    ]
    ws = wb.create_sheet("পুরনো+নতুন মিলেছে")
    ws.append(MERGED_MATCH_HEADERS)
    for match_no, record in enumerate(matched_records, start=1):
        phones = record["phones"]
        old_phones = record.get("_matched_old_phones") or []
        ws.append([
            match_no,
            record.get("_match_type", ""),
            record.get("_clean_row", ""),
            record.get("source_row", ""),
            record.get("_matched_old_row", ""),
            record["jamat"],
            record["permanent_id"],
            record["name"],
            record.get("_matched_old_name", ""),
            record["father_name"],
            "; ".join(phones),
            "; ".join(old_phones),
            record["study_detail"],
            record.get("_matched_old_study_detail", ""),
            " | ".join(dict.fromkeys([c for c in record["comments"] if c])),
            record["source"],
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{max(ws.max_row, 1)}"
    style_header(ws[1], header_fill, header_font, border)
    style_body(ws, border)
    widths = {
        "A": 12, "B": 18, "C": 12, "D": 14, "E": 14, "F": 14, "G": 16,
        "H": 28, "I": 28, "J": 30, "K": 24, "L": 24, "M": 26, "N": 32,
        "O": 54, "P": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 42
    if ws.max_row >= 2:
        table = Table(displayName="MergedMatchTable", ref=f"A1:P{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium6",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws.sheet_view.showGridLines = False
    return len(matched_records)


def write_workbook(records: list[dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Alumni List"

    ws.append(HEADERS)
    for clean_row, record in enumerate(records, start=2):
        record["_clean_row"] = clean_row
        ws.append(flatten_row(record))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{ws.max_row}"

    header_fill = PatternFill("solid", fgColor="1F4E3D")
    header_font = Font(color="FFFFFF", bold=True, name="Nirmala UI")
    thin = Side(style="thin", color="D9E2D7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    style_header(ws[1], header_fill, header_font, border)
    style_body(ws, border)

    widths = {
        "A": 14, "B": 16, "C": 28, "D": 30, "E": 14, "F": 18, "G": 38,
        "H": 44, "I": 15, "J": 15, "K": 22, "L": 16, "M": 16, "N": 26,
        "O": 14, "P": 50, "Q": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 42

    status_validation = DataValidation(type="list", formula1='"সম্পূর্ণ,অসম্পূর্ণ"', allow_blank=True)
    status_validation.error = "শুধু সম্পূর্ণ অথবা অসম্পূর্ণ নির্বাচন করুন"
    status_validation.prompt = "ম্যানুয়ালি নির্বাচন করুন"
    ws.add_data_validation(status_validation)
    status_validation.add(f"O2:O{ws.max_row}")

    table = Table(displayName="CleanAlumniTable", ref=f"A1:Q{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    ws.sheet_view.showGridLines = False
    red_count, yellow_count = apply_review_fills(ws, records)
    duplicate_group_count = write_duplicate_sheet(wb, records, header_fill, header_font, border)
    wb.properties.comments = f"red_review_rows={red_count}; yellow_review_rows={yellow_count}; possible_duplicate_groups={duplicate_group_count}"
    wb.save(OUT_FILE)


def main() -> None:
    old_records = parse_old_records()
    new_records = parse_new_records()
    combined = combine_records(old_records, new_records)
    write_workbook(combined)
    print(f"old_records={len(old_records)}")
    print(f"new_records={len(new_records)}")
    print(f"combined_rows={len(combined)}")
    print(f"output={OUT_FILE}")


if __name__ == "__main__":
    main()
